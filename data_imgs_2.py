# Importing Libraries
from PIL import Image
from PIL.ExifTags import TAGS
import glob
import os
from openpyxl import Workbook, load_workbook
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl import formatting, styles
import pandas as pd

# Taking all images files from folder
images = glob.glob("*.jpg")
result = []
for image in images:
    # Path to the image 
    imagename = image
    absolute_path = os.path.abspath(image)
    # Read the image data using PIL
    image = Image.open(imagename)
    # Extracting data
    info_dict = [image.filename,image.format,absolute_path,str(image.size), str(image.height), str(image.width)]
    result.append(info_dict)

# Saving to excel
wb = openpyxl.Workbook()
ws1 = wb.active
# Inserting headers
headers = ["Image Name", "Image Extension", "Full Path", "Image Size", "Image Height", "Image Width"]  
ws1.append(headers)
# Inserting values 
for row in result:   
    ws1.append(row)
wb.save("Images_Data.xlsx")
# Changing the font and color of headers
light_green_font = Font(color='00FF00', italic=True)
for cell in ws1["1:1"]:
    cell.font = light_green_font



# df = pd.read_excel("Images_Data.xlsx")
# data = df.loc[(df['Image Height'] >= 1000) & (df['Image Width'] >= 1000)]
# print(data)

# Iterating on columns to get the cell
a = []
for col1 in ws1['E']:
    if len(col1.value) > 3:
        col1.value
        #print(col1.value)
        a.append(col1.value)
b = []
for col2 in ws1['F']:
    if len(col2.value) > 3:
        col2.value
        #print(col2.value)
        b.append(col2.value)
# a = col1.value
# b = col2.value
print(a)
print(b)


# my_fill = openpyxl.styles.colors.Color(rgb='00FF0000')
# my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_fill)
# data.fill = my_fill


wb.save("Images_Data_2.xlsx")