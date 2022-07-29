# Importing Libraries
from PIL import Image
from PIL.ExifTags import TAGS
import glob
import os
from openpyxl import Workbook, load_workbook
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills

# Taking all images files from folder
images = glob.glob("*.jpg")
result = []

for image in images:
    # Path to the image 
    imagename = image
    absolute_path = os.path.abspath(image)
    # Read the image data using PIL
    image = Image.open(imagename)
    # Extracting data
    info_dict = [image.filename,image.format,absolute_path,str(image.size), str(image.height), str(image.width)]
    result.append(info_dict)

# Saving to excel
wb = openpyxl.Workbook()
ws1 = wb.active
# Inserting headers
headers = ["Image Name", "Image Extension", "Full Path", "Image Size", "Image Height", "Image Width"]  
ws1.append(headers)
# Inserting values 
for row in result:   
    ws1.append(row)

# Changing the font and color of images 
red_font = Font(color='FF0000', italic=True)
for col1 in ws1['E']:
    if len(col1.value) > 3:
        col1.font=red_font
            
for col2 in ws1['F']:
    if len(col2.value) > 3:
        col2.font=red_font
        
# Changing the font and color of header
blue_font = Font(color='0000EE', italic=True)
for cell in ws1["1:1"]:
    cell.font = blue_font

wb.save("Images_Data.xlsx")